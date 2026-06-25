"""将业务 Excel 导出为运行时使用的版本化需求词 JSON。"""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


PRIORITY_MAP = {
    "一级词": "level_1",
    "二级词": "level_2",
    "三级热点词": "level_3",
}


def export_catalog(source: Path, destination: Path) -> None:
    workbook = load_workbook(source, data_only=True)
    game_sheet = workbook["全游戏需求词总表"]
    generic_sheet = workbook["通用监控词总表"]

    games: dict[str, list[dict]] = {}
    for game_name, category, priority, term in game_sheet.iter_rows(
        min_row=3,
        max_col=4,
        values_only=True,
    ):
        if not game_name or not term:
            continue
        games.setdefault(str(game_name).strip(), []).append({
            "category": str(category).strip(),
            "priority": PRIORITY_MAP[str(priority).strip()],
            "term": str(term).strip(),
        })

    generic: list[dict] = []
    for category, priority, term in generic_sheet.iter_rows(
        min_row=3,
        max_col=3,
        values_only=True,
    ):
        if not term:
            continue
        generic.append({
            "category": str(category).strip(),
            "priority": PRIORITY_MAP[str(priority).strip()],
            "term": str(term).strip(),
        })

    payload = {
        "source": source.name,
        "version": "2026-06-25",
        "games": games,
        "generic": generic,
    }
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "app" / "data" / "demand_keywords.json",
    )
    args = parser.parse_args()
    export_catalog(args.source, args.output)


if __name__ == "__main__":
    main()
